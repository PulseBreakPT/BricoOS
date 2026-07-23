"""Anexos_Edição.zip do Correio Semanal — extração determinística de um
pacote de ficheiros heterogéneo (Excel, PDF, e por vezes um zip dentro do
zip), sem depender do nome dos ficheiros nem de uma estrutura fixa de
pastas: cada ficheiro é aberto e classificado pelo próprio conteúdo.

Decisão de âmbito (confirmada): motor determinístico, não IA — deteta o
que reconhecer por palavra-chave (cabeçalhos de coluna em Excel, termos no
texto de um PDF); um layout genuinely novo, sem nenhuma palavra-chave
conhecida, fica arquivado e classificado como "outro" em vez de ficar sem
dados — nunca falha nem é ignorado, só não é decomposto em campos.

A classificação por categoria e a deteção de factos (EAN, preços,
contactos, ...) usam os motores partilhados classification_rules.py e
extraction_patterns.py — a mesma lógica que serve o correio_semanal.py,
para o "aprender sozinho" (categorias novas guardadas na base de dados,
nunca no código) funcionar da mesma forma nos dois lados.
"""

import hashlib
import io
import re
import unicodedata
import zipfile
from collections import Counter

import fitz  # PyMuPDF
import openpyxl

try:
    import classification_rules as clsrules
    import extraction_patterns
    import severity as severity_engine
except ImportError:  # Permite também executar como módulo: python -m backend.server
    from . import classification_rules as clsrules
    from . import extraction_patterns
    from . import severity as severity_engine

MAX_ROWS_PER_SHEET = 1000
HEADER_SCAN_ROWS = 15
MIN_TEXT_CHARS_PER_PAGE = 40  # abaixo disto, o PDF é provavelmente imagem/vetor sem texto


def _fold(text):
    normalized = unicodedata.normalize("NFKD", text or "")
    return "".join(c for c in normalized if not unicodedata.combining(c)).lower()


def _content_hash(data):
    return hashlib.sha256(data).hexdigest()


def iter_zip_files(zip_bytes, _prefix=""):
    """Gerador (caminho, bytes) para cada ficheiro do zip — desce
    recursivamente em qualquer zip encontrado dentro do zip, sem limite de
    profundidade fixo (o padrão real: um "Anexos_Edição.zip" que contém um
    "AÇÃO_...zip" com mais PDFs lá dentro)."""
    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except zipfile.BadZipFile:
        return
    for info in zf.infolist():
        if info.is_dir():
            continue
        try:
            data = zf.read(info)
        except Exception:
            continue
        path = _prefix + info.filename
        if path.lower().endswith(".zip"):
            yield from iter_zip_files(data, _prefix=path + "/")
        else:
            yield path, data


_EXT_KIND = {
    ".xlsx": "xlsx", ".xlsm": "xlsx", ".xls": "xlsx",
    ".pdf": "pdf", ".docx": "docx", ".doc": "docx",
    ".jpg": "image", ".jpeg": "image", ".png": "image", ".gif": "image",
}


def detect_file_kind(path, data):
    ext = "." + path.rsplit(".", 1)[-1].lower() if "." in path.rsplit("/", 1)[-1] else ""
    kind = _EXT_KIND.get(ext)
    if kind:
        return kind
    if data[:4] == b"%PDF":
        return "pdf"
    if data[:2] == b"PK":
        return "xlsx"  # zip-based Office format sem extensão reconhecida
    return "other"


# ---------- Classificação por conteúdo (nunca pelo nome do ficheiro) ----------
# As categorias e as palavras-chave já não vivem aqui — ver
# classification_rules.py: ficam na base de dados, partilhadas com o
# correio_semanal.py, e podem crescer sem alterar código nenhum (ver
# classification_rules.learn_new_category). `rules` é sempre o dicionário
# devolvido por classification_rules.load_rules(db), carregado uma vez por
# email em server.py e passado como parâmetro até aqui.


# ---------- Excel: deteção flexível de cabeçalho (por palavra-chave, não posição) ----------
# Cada campo canónico tem um conjunto de padrões (regex, já sobre texto sem
# acentos) que uma coluna de cabeçalho pode ter — várias colunas podem
# mapear para o mesmo campo (ex.: duas colunas de preço = preço antigo e
# novo lado a lado, muito comum nestas tabelas).
_FIELD_PATTERNS = {
    "ean": (r"^ean", r"^ean.?13$"),
    "itm": (r"^itm", r"^codigo", r"^cod\.?$", r"^ref\.?$", r"^referencia"),
    "descricao": (r"description", r"descri[cç][aã]o", r"descripci[oó]n", r"designa[cç][aã]o", r"^produto$"),
    "marca": (r"^marca$", r"^fornecedor$", r"rais.?soc", r"^marque$"),
    "preco": (r"pre[cç]o", r"precio", r"^pvp", r"tarifa", r"^p\.?\s*cess", r"^pc\b"),
    "quantidade": (r"^uds", r"quantidade", r"^cdt$", r"acond"),
    "data": (r"^data", r"^fecha", r"vigor", r"in[ií]cio", r"^fim$"),
    "incidente": (r"incidente", r"causa"),
    "plano_accao": (r"plano.?a[cç][cç][aã]o", r"solu[cç][aã]o"),
    "estado": (r"^estado$", r"^status$"),
}
_FIELD_RES = {field: [re.compile(p) for p in pats] for field, pats in _FIELD_PATTERNS.items()}


def _match_fields(header_text):
    folded = _fold(header_text).strip()
    return [field for field, res in _FIELD_RES.items() if any(r.search(folded) for r in res)]


_MAX_COLS = 40


def _extract_sheet(ws):
    """openpyxl em modo read-only não é fiável em ws.max_row/max_column —
    ficheiros reais aparecem com formatação aplicada à coluna inteira e o
    valor fica ordens de grandeza acima do conteúdo real (chegou a
    pendurar o processo minutos a fio numa das folhas de amostra). Por
    isso nunca se itera por posição: só ws.iter_rows() em stream, com um
    limite explícito de linhas (MAX_ROWS_PER_SHEET) que manda parar
    independentemente do que os metadados da folha disserem."""
    row_iter = ws.iter_rows(values_only=True)
    buffered = []
    for _ in range(HEADER_SCAN_ROWS):
        try:
            buffered.append(next(row_iter))
        except StopIteration:
            break

    best_row, best_score, best_fields = None, 0, {}
    for idx, values in enumerate(buffered, start=1):
        fields, score = {}, 0
        for c, v in enumerate(values[:_MAX_COLS], start=1):
            if not v or not isinstance(v, str):
                continue
            matched = _match_fields(v)
            if matched:
                fields[c] = {"header": v.strip(), "fields": matched}
                score += 1
        if score > best_score:
            best_row, best_score, best_fields = idx, score, fields

    if not best_row or not best_fields:
        # Sem cabeçalho reconhecível — não inventa colunas; guarda só uma
        # amostra do texto para a categoria poder ser deduzida de outra forma.
        sample = " ".join(str(v) for row in buffered for v in row if isinstance(v, str))[:1000]
        return {"header_row": None, "columns": {}, "rows": [], "row_count": 0,
                "truncated": False, "sample_text": sample}

    # Amostra para a classificação por conteúdo: todo o texto das linhas até
    # ao cabeçalho (título/contexto acima) + a linha de cabeçalho inteira —
    # não só as colunas reconhecidas, senão uma coluna como "Tabela 2026
    # c/sobretaxa" perdia a palavra "sobretaxa" por a coluna em si não
    # corresponder a nenhum campo canónico.
    context_rows = buffered[:best_row]
    sample = " ".join(str(v) for row in context_rows for v in row if isinstance(v, str))[:1500]

    columns = best_fields
    rows, truncated = [], False

    def _consume(values):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            return
        row = {}
        for c, v in enumerate(values[:_MAX_COLS], start=1):
            if c in columns and v is not None:
                row[columns[c]["header"]] = v
        if row:
            rows.append(row)

    for values in buffered[best_row:]:
        if len(rows) >= MAX_ROWS_PER_SHEET:
            truncated = True
            break
        _consume(values)
    if not truncated:
        for values in row_iter:
            if len(rows) >= MAX_ROWS_PER_SHEET:
                truncated = True
                break
            _consume(values)

    return {"header_row": best_row, "columns": {c: v["fields"] for c, v in columns.items()},
            "column_labels": {c: v["header"] for c, v in columns.items()},
            "rows": rows, "row_count": len(rows), "truncated": truncated, "sample_text": sample}


def _structural_category(sheets):
    """As frases-chave no texto às vezes não chegam (uma folha só com
    "EAN / DESIGNAÇÃO / PREÇO 2026" não usa nenhuma das frases da lista),
    mas as colunas já identificadas por _extract_sheet não deixam dúvidas
    sobre o que a folha é — mais fiável do que procurar frases soltas."""
    all_fields = {f for sheet in sheets for cols in sheet.get("columns", {}).values() for f in cols}
    if "incidente" in all_fields or "plano_accao" in all_fields:
        return "incidencia"
    if "preco" in all_fields and ("ean" in all_fields or "itm" in all_fields or "descricao" in all_fields):
        return "tabela_precos"
    return None


def extract_xlsx(data, rules):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        return {"error": str(e), "sheets": []}
    sheets = []
    classify_text = []
    for ws in wb.worksheets:
        info = _extract_sheet(ws)
        info["sheet_name"] = ws.title
        sheets.append(info)
        classify_text.append(ws.title)
        classify_text.append(info.get("sample_text", ""))
    joined = " ".join(classify_text)
    category = clsrules.classify(joined, rules)
    if category == "outro":
        category = _structural_category(sheets) or category
    facts = extraction_patterns.extract_facts(joined)
    result = {"sheets": sheets, "category": category}
    if facts:
        result["facts"] = facts
    if category == "outro":
        result["fingerprint"] = clsrules.unclassified_fingerprint(joined)
    return result


# ---------- PDF: reaproveita a mesma extração de texto de correio_semanal.py ----------
_FACT_CONTACT_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+|\b2\d{2}\s?\d{3}\s?\d{3}\b")
_FACT_DEADLINE_RE = re.compile(r"DATA LIMITE DE (RESPOSTA|ENCOMENDA)[:_\s]*", re.IGNORECASE)
_ADHESION_RE = re.compile(r"N[ÃA]O\s+ADES[ÃA]O|ADES[ÃA]O", re.IGNORECASE)


def extract_pdf(data, rules):
    try:
        doc = fitz.open(stream=data, filetype="pdf")
    except Exception as e:
        return {"error": str(e)}
    page_count = doc.page_count
    texts = [page.get_text() for page in doc]
    full_text = "\n".join(texts)
    avg_chars = len(full_text) / max(page_count, 1)
    if avg_chars < MIN_TEXT_CHARS_PER_PAGE:
        return {"page_count": page_count, "likely_image_pdf": True, "category": "outro",
                "text_chars": len(full_text)}
    category = clsrules.classify(full_text, rules)
    deadlines = sorted(set(m.group(0).strip() for m in _FACT_DEADLINE_RE.finditer(full_text)))
    contacts = sorted(set(_FACT_CONTACT_RE.findall(full_text)))
    result = {"page_count": page_count, "likely_image_pdf": False, "category": category,
              "text_chars": len(full_text), "deadline_mentions": deadlines, "contacts": contacts,
              "excerpt": re.sub(r"\s+", " ", full_text).strip()[:600]}
    facts = extraction_patterns.extract_facts(full_text)
    if facts:
        result["facts"] = facts
    if category == "outro":
        result["fingerprint"] = clsrules.unclassified_fingerprint(full_text)
    return result


_EDITION_RE = re.compile(r"(\d{3,4})")


def process_zip(zip_bytes, zip_filename="", rules=None):
    """Ponto de entrada: descompacta (incl. zips aninhados), classifica e
    extrai cada ficheiro pelo conteúdo. Devolve a lista de ficheiros
    processados + um resumo — nunca lança exceção por um ficheiro
    individual falhar (fica marcado com "error", os restantes continuam).

    `rules` é o dicionário de categorias/palavras-chave carregado da base
    de dados (classification_rules.load_rules) — se não for passado, usa
    a semente predefinida (útil para testes isolados deste módulo)."""
    rules = rules if rules is not None else clsrules.DEFAULT_KEYWORDS
    edition_m = _EDITION_RE.search(zip_filename or "")
    edition_number = edition_m.group(1) if edition_m else None

    files = []
    for path, data in iter_zip_files(zip_bytes):
        entry = {"path": path, "size": len(data), "content_hash": _content_hash(data),
                 "kind": detect_file_kind(path, data)}
        try:
            if entry["kind"] == "xlsx":
                extracted = extract_xlsx(data, rules)
                entry["category"] = extracted.pop("category", "outro")
                entry["extracted"] = extracted
            elif entry["kind"] == "pdf":
                extracted = extract_pdf(data, rules)
                entry["category"] = extracted.get("category", "outro")
                entry["extracted"] = extracted
            else:
                entry["category"] = "outro"
                entry["extracted"] = {}
        except Exception as e:
            entry["category"] = "outro"
            entry["extracted"] = {}
            entry["error"] = str(e)
        if entry["kind"] == "xlsx":
            sample_text = " ".join(s.get("sample_text", "") for s in entry["extracted"].get("sheets", []))
        else:
            sample_text = entry["extracted"].get("excerpt") or ""
        requires_adhesion = bool(_ADHESION_RE.search(sample_text)) if sample_text else False
        entry["severity"] = severity_engine.classify_severity(
            entry["category"], sample_text, requires_adhesion=requires_adhesion)
        files.append(entry)

    by_category = Counter(f["category"] for f in files)
    by_severity = Counter(f["severity"] for f in files)
    errors = sum(1 for f in files if f.get("error"))
    unclassified = [{"path": f["path"], "fingerprint": f["extracted"]["fingerprint"]}
                     for f in files if f["category"] == "outro" and f["extracted"].get("fingerprint")]
    return {
        "edition_number": edition_number,
        "zip_filename": zip_filename,
        "file_count": len(files),
        "files": files,
        "by_category": dict(by_category),
        "by_severity": dict(by_severity),
        "error_count": errors,
        "unclassified_fingerprints": unclassified,
    }


# ---------- Comparação entre edições ----------
def price_index(processed):
    """{ean_ou_itm: {"preco", "descricao", "path"}} — juntando TODAS as
    tabelas de preços da edição, não um ficheiro em concreto: os nomes dos
    ficheiros de preços mudam de semana para semana, o que se mantém
    comparável é o conteúdo (EAN/ITM). Quando uma folha tem mais do que
    uma coluna de preço lado a lado (ex.: "Tabela 2025" e "Tabela 2026" —
    antigo vs novo dentro da própria semana), usa-se a mais à direita, que
    nas amostras reais é sempre a mais recente."""
    index = {}
    for f in processed.get("files", []):
        if f.get("category") != "tabela_precos" or f.get("kind") != "xlsx":
            continue
        for sheet in f.get("extracted", {}).get("sheets", []):
            columns, labels = sheet.get("columns", {}), sheet.get("column_labels", {})
            key_cols = [c for c, fs in columns.items() if "ean" in fs or "itm" in fs]
            desc_cols = [c for c, fs in columns.items() if "descricao" in fs]
            preco_cols = sorted(c for c, fs in columns.items() if "preco" in fs)
            if not key_cols or not preco_cols:
                continue
            key_label = labels[key_cols[0]]
            desc_label = labels[desc_cols[0]] if desc_cols else None
            preco_label = labels[preco_cols[-1]]
            for row in sheet.get("rows", []):
                key, price = row.get(key_label), row.get(preco_label)
                if key is None or not isinstance(price, (int, float)):
                    continue
                index[str(key)] = {"preco": float(price),
                                    "descricao": row.get(desc_label) if desc_label else None,
                                    "path": f["path"]}
    return index


def diff_edicao_versions(prev, new):
    """Compara duas edições do Anexos_Edição — ficheiros novos/removidos/
    atualizados (por hash de conteúdo, nunca pelo nome — os nomes mudam de
    semana para semana) e alterações de preço/produtos novos/descontinuados
    (por EAN/ITM, juntando todas as tabelas de preços de cada edição)."""
    if not prev or not new:
        return None

    def norm(path):
        # O primeiro segmento do caminho é a pasta raiz do zip
        # ("Anexos_Edição 644/...") — muda de número a cada edição, por
        # isso não entra na comparação.
        parts = path.split("/", 1)
        return parts[1] if len(parts) > 1 else path

    prev_by_path = {norm(f["path"]): f for f in prev.get("files", [])}
    new_by_path = {norm(f["path"]): f for f in new.get("files", [])}
    new_files = sorted(set(new_by_path) - set(prev_by_path))
    removed_files = sorted(set(prev_by_path) - set(new_by_path))
    updated_files = sorted(
        p for p in (set(new_by_path) & set(prev_by_path))
        if new_by_path[p]["content_hash"] != prev_by_path[p]["content_hash"])

    prev_prices, new_prices = price_index(prev), price_index(new)
    price_changes = []
    for key, new_info in new_prices.items():
        old_info = prev_prices.get(key)
        if not old_info or abs(old_info["preco"] - new_info["preco"]) < 0.005:
            continue
        pct = ((new_info["preco"] - old_info["preco"]) / old_info["preco"] * 100) if old_info["preco"] else None
        price_changes.append({
            "key": key, "descricao": new_info.get("descricao") or old_info.get("descricao"),
            "preco_antigo": old_info["preco"], "preco_novo": new_info["preco"],
            "variacao_pct": round(pct, 1) if pct is not None else None})
    price_changes.sort(key=lambda c: -(abs(c["variacao_pct"]) if c["variacao_pct"] is not None else 0))
    new_products = sorted(set(new_prices) - set(prev_prices))
    discontinued_products = sorted(set(prev_prices) - set(new_prices))

    return {
        "prev_edition": prev.get("edition_number"), "new_edition": new.get("edition_number"),
        "new_files": new_files, "removed_files": removed_files, "updated_files": updated_files,
        "price_changes": price_changes[:200], "price_changes_count": len(price_changes),
        "new_products": new_products[:200], "new_products_count": len(new_products),
        "discontinued_products": discontinued_products[:200],
        "discontinued_products_count": len(discontinued_products),
        "has_changes": bool(new_files or removed_files or updated_files or price_changes
                            or new_products or discontinued_products),
    }
